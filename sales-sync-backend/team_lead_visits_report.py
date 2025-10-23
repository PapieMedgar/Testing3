import csv
import os
import sys
import glob
from datetime import date, timedelta
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


TEAM_LEAD_TO_USERS: Dict[str, List[str]] = {
    "068 641 1128": [
        "069 067 6463",
        "069 069 8934",
        "068 617 5687",
        "068 616 1031",
        "068 641 1128",
    ],
    "069 043 3247": [
        "063 901 9701",
        "069 027 3894",
        "068 907 8688",
        "069 034 4500",
        "069 033 8517",
        "069 043 3247",
    ],
    "069 068 2819": [
        "068 639 0928",
        "069 067 7526",
        "069 027 9986",
        "069 062 5752",
        "069 052 7867",
        "069 068 2819",
    ],
    "069 066 2955": [
        "068 609 0618",
        "063 905 0286",
        "069 058 4391",
        "069 069 7103",
        "068 602 7701",
        "069 066 2955",
    ],
}


LEAD_ORDER: List[str] = [
    "068 641 1128",
    "069 043 3247",
    "069 068 2819",
    "069 066 2955",
]


def find_latest_daily_visits_file(reports_dir: str = "reports") -> str:
    """Find the most recent daily visits CSV file."""
   
    pattern = os.path.join(reports_dir, "daily_visits_*.csv")
    files = glob.glob(pattern)
    
    if not files:
       
        old_file = os.path.join(reports_dir, "daily_visits.csv")
        if os.path.exists(old_file):
            return old_file
        return None
    
   
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0]


def generate_daily_visits_if_needed(reports_dir: str = "reports") -> str:
    """Generate daily visits report if it doesn't exist."""
    
    latest_file = find_latest_daily_visits_file(reports_dir)
    if latest_file and os.path.exists(latest_file):
       
        file_time = os.path.getmtime(latest_file)
        current_time = os.path.getmtime(__file__)  
        if current_time - file_time < 7 * 24 * 3600: 
            return latest_file
    
    
    print("Generating daily visits report...")
    try:
        from daily_visits_report import main as generate_daily_visits
        
        yesterday = date.today() - timedelta(days=1)
        date_str = yesterday.strftime('%Y-%m-%d')
        output_file = os.path.join(reports_dir, f"daily_visits_{date_str}.csv")
        
        
        original_argv = sys.argv
        sys.argv = ['daily_visits_report.py', date_str, date_str, output_file]
        
        try:
            generate_daily_visits()
            return output_file
        finally:
            sys.argv = original_argv
            
    except Exception as e:
        print(f"Warning: Could not auto-generate daily visits report: {e}")
        return None


def read_daily_visits_pivot(input_csv_path: str) -> List[Dict[str, int]]:
    """Read the existing per-user daily visits pivot CSV.

    Returns a list of dictionaries where each dict has keys:
      - "Date": date string as-is from the CSV (e.g., 01-Oct-25)
      - other keys are user display names with integer visit counts
    """
    if not os.path.exists(input_csv_path):
        print(f"Error: Input file not found: {input_csv_path}")
        print("This script requires a daily visits report as input.")
        print("Please run one of the following commands first:")
        print("  python Scripts/daily_visits_report.py")
        print("  python Scripts/daily_visits_report.py 2024-01-15 2024-01-15 reports/daily_visits_2024-01-15.csv")
        print("Or specify the correct input file as an argument:")
        print("  python Scripts/team_lead_visits_report.py path/to/daily_visits.csv")
        raise SystemExit(1)

    rows: List[Dict[str, int]] = []
    with open(input_csv_path, mode="r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for raw in reader:
            parsed: Dict[str, int] = {"Date": raw.get("Date", "")}
            for key, val in raw.items():
                if key == "Date":
                    continue
                try:
                    parsed[key] = int(val) if val not in (None, "", " ") else 0
                except Exception:
                    parsed[key] = 0
            rows.append(parsed)
    return rows


essential_users: List[str] = sorted({u for users in TEAM_LEAD_TO_USERS.values() for u in users})


def compute_team_lead_sums(
    per_user_rows: List[Dict[str, int]], mapping: Dict[str, List[str]]
) -> List[Dict[str, int]]:
    """Aggregate per-user counts to per-team-lead sums per date.

    Missing users are treated as 0.
    """
    aggregated: List[Dict[str, int]] = []
    for row in per_user_rows:
        out_row: Dict[str, int] = {"Date": row["Date"]}
        for lead, users in mapping.items():
            total = 0
            for user in users:
                total += int(row.get(user, 0) or 0)
            out_row[lead] = total
        aggregated.append(out_row)
    return aggregated


def write_team_lead_csv(rows: List[Dict[str, int]], output_csv_path: str) -> None:
    os.makedirs(os.path.dirname(output_csv_path), exist_ok=True)
    headers = ["Date"] + LEAD_ORDER + ["Total"]
    with open(output_csv_path, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for r in rows:
            lead_values = [int(r.get(lead, 0) or 0) for lead in LEAD_ORDER]
            total_value = sum(lead_values)
            row_out = {"Date": r.get("Date", "")}
            for lead, value in zip(LEAD_ORDER, lead_values):
                row_out[lead] = value
            row_out["Total"] = total_value
            writer.writerow(row_out)


def write_team_lead_xlsx(rows: List[Dict[str, int]], output_csv_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Visits by Team Lead"

    headers = ["Date"] + LEAD_ORDER + ["Total"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        lead_values = [int(r.get(lead, 0) or 0) for lead in LEAD_ORDER]
        total_value = sum(lead_values)
        ws.append([r.get("Date", "")] + lead_values + [total_value])

    # Autosize columns
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    xlsx_path = output_csv_path.replace(".csv", ".xlsx")
    wb.save(xlsx_path)
    return xlsx_path


def generate_xlsx_bytes(start_date=None, end_date=None):
    """Generate the team lead visits Excel file as bytes for API download, always fresh."""
    try:
        import csv
        from io import StringIO
        from daily_visits_report import generate_csv_bytes as generate_daily_visits_csv
        
        csv_bytes = generate_daily_visits_csv(start_date, end_date)
        csv_str = csv_bytes.decode("utf-8")
        reader = csv.DictReader(StringIO(csv_str))
        per_user_rows = []
        for raw in reader:
            parsed = {"Date": raw.get("Date", "")}
            for key, val in raw.items():
                if key == "Date":
                    continue
                try:
                    parsed[key] = int(val) if val not in (None, "", " ") else 0
                except Exception:
                    parsed[key] = 0
            per_user_rows.append(parsed)
        aggregated_rows = compute_team_lead_sums(per_user_rows, TEAM_LEAD_TO_USERS)
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Visits by Team Lead"
        headers = ["Date"] + LEAD_ORDER + ["Total"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for r in aggregated_rows:
            lead_values = [int(r.get(lead, 0) or 0) for lead in LEAD_ORDER]
            total_value = sum(lead_values)
            ws.append([r.get("Date", "")] + lead_values + [total_value])
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        from io import BytesIO
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.getvalue()
    except Exception as e:
        import traceback
        print('Error in generate_xlsx_bytes:', e)
        traceback.print_exc()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Error"
        ws.append(["Failed to generate report", str(e)])
        from io import BytesIO
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.getvalue()


def generate_csv_bytes(start_date=None, end_date=None):
    """
    Generate the team lead visits CSV as bytes for API download (in-memory, no file I/O).
    """
    import traceback
    from io import StringIO
    try:
        import csv
        from daily_visits_report import generate_csv_bytes as generate_daily_visits_csv
        
        csv_bytes = generate_daily_visits_csv(start_date, end_date)
        csv_str = csv_bytes.decode("utf-8")
        reader = csv.DictReader(StringIO(csv_str))
        per_user_rows = []
        for raw in reader:
            parsed = {"Date": raw.get("Date", "")}
            for key, val in raw.items():
                if key == "Date":
                    continue
                try:
                    parsed[key] = int(val) if val not in (None, "", " ") else 0
                except Exception:
                    parsed[key] = 0
            per_user_rows.append(parsed)
        aggregated_rows = compute_team_lead_sums(per_user_rows, TEAM_LEAD_TO_USERS)
        headers = ["Date"] + LEAD_ORDER + ["Total"]
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        for r in aggregated_rows:
            lead_values = [int(r.get(lead, 0) or 0) for lead in LEAD_ORDER]
            total_value = sum(lead_values)
            row_out = {"Date": r.get("Date", "")}
            for lead, value in zip(LEAD_ORDER, lead_values):
                row_out[lead] = value
            row_out["Total"] = total_value
            writer.writerow(row_out)
        return output.getvalue().encode("utf-8")
    except Exception as e:
        print('Error in generate_csv_bytes:', e)
        traceback.print_exc()
        
        return f"Error,Failed to generate report: {e}\n".encode("utf-8")


def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(description="Generate team lead visits report for a date range.")
    parser.add_argument("--start_date", type=str, default=None, help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end_date", type=str, default=None, help="End date (YYYY-MM-DD)")
    args = parser.parse_args()

    
    from daily_visits_report import generate_csv_bytes as generate_daily_visits_csv
    import csv
    from io import StringIO
    csv_bytes = generate_daily_visits_csv(args.start_date, args.end_date)
    csv_str = csv_bytes.decode("utf-8")
    reader = csv.DictReader(StringIO(csv_str))
    per_user_rows = []
    for raw in reader:
        parsed = {"Date": raw.get("Date", "")}
        for key, val in raw.items():
            if key == "Date":
                continue
            try:
                parsed[key] = int(val) if val not in (None, "", " ") else 0
            except Exception:
                parsed[key] = 0
        per_user_rows.append(parsed)
    aggregated_rows = compute_team_lead_sums(per_user_rows, TEAM_LEAD_TO_USERS)
    headers = ["Date"] + LEAD_ORDER + ["Total"]
   
    print("\t".join(headers))
    for r in aggregated_rows:
        lead_values = [int(r.get(lead, 0) or 0) for lead in LEAD_ORDER]
        total_value = sum(lead_values)
        row = [r.get("Date", "")] + [str(v) for v in lead_values] + [str(total_value)]
        print("\t".join(row))


if __name__ == "__main__":
    main()