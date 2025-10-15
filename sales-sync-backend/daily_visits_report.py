import csv
import os
import sys
from datetime import datetime, date
from typing import Optional
from io import BytesIO

import mysql.connector
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

try:
    sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
    from db_config import DATABASE_CONFIG
except Exception as import_error:
    print(f"Failed to import DATABASE_CONFIG: {import_error}")
    sys.exit(1)


def parse_date(date_str: Optional[str]) -> Optional[date]:
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError as exc:
        raise SystemExit(f"Invalid date format '{date_str}'. Use YYYY-MM-DD.") from exc


def build_query(
    start_date: Optional[date],
    end_date: Optional[date],
    *,
    users_table: str,
    users_pk: str,
    users_name_expr: str,
    checkins_table: str,
    checkins_pk: str,
    checkins_user_id_col: str,
    checkins_time_col: str,
    include_visit_response: bool,
    visit_response_table: str,
    vr_checkin_id_col: str,
) -> tuple[str, list]:
    where_clauses = ["1=1"]
    params: list = []

    if start_date is not None:
        where_clauses.append(f"DATE(c.`{checkins_time_col}`) >= %s")
        params.append(start_date.isoformat())
    if end_date is not None:
        where_clauses.append(f"DATE(c.`{checkins_time_col}`) <= %s")
        params.append(end_date.isoformat())

    where_sql = " AND ".join(where_clauses)

    vr_join = (
        f" LEFT JOIN `{visit_response_table}` vr ON vr.`{vr_checkin_id_col}` = c.`{checkins_pk}`\n"
        if include_visit_response
        else ""
    )

    query = f"""
        SELECT
            {users_name_expr} AS user_name,
            DATE(c.`{checkins_time_col}`) AS visit_date,
            COUNT(c.`{checkins_pk}`) AS total_visits
        FROM `{users_table}` u
        JOIN `{checkins_table}` c ON c.`{checkins_user_id_col}` = u.`{users_pk}`
        {vr_join}
        WHERE {where_sql}
        GROUP BY {users_name_expr}, DATE(c.`{checkins_time_col}`)
        ORDER BY {users_name_expr} ASC, visit_date ASC
    """
    return query, params


def table_exists(connection, table_name: str) -> bool:
    cursor = connection.cursor()
    try:
        cursor.execute(
            """
            SELECT COUNT(*)
            FROM information_schema.tables
            WHERE table_schema = DATABASE() AND table_name = %s
            """,
            (table_name,),
        )
        count = cursor.fetchone()[0]
        return bool(count)
    finally:
        cursor.close()


def resolve_table_name(connection, candidates: list[str]) -> Optional[str]:
    for name in [c for c in candidates if c]:
        if table_exists(connection, name):
            return name
    return None


def get_primary_key_column(connection, table_name: str) -> Optional[str]:
    cursor = connection.cursor()
    try:
        cursor.execute(
            """
            SELECT k.COLUMN_NAME
            FROM information_schema.table_constraints t
            JOIN information_schema.key_column_usage k
              ON t.CONSTRAINT_NAME = k.CONSTRAINT_NAME
             AND t.TABLE_SCHEMA = k.TABLE_SCHEMA
             AND t.TABLE_NAME = k.TABLE_NAME
            WHERE t.TABLE_SCHEMA = DATABASE()
              AND t.TABLE_NAME = %s
              AND t.CONSTRAINT_TYPE = 'PRIMARY KEY'
            ORDER BY k.ORDINAL_POSITION
            LIMIT 1
            """,
            (table_name,),
        )
        row = cursor.fetchone()
        return row[0] if row else None
    finally:
        cursor.close()


def get_columns_for_table(connection, table_name: str) -> list[tuple[str, str]]:
    cursor = connection.cursor()
    try:
        cursor.execute(
            """
            SELECT COLUMN_NAME, DATA_TYPE
            FROM information_schema.columns
            WHERE table_schema = DATABASE() AND table_name = %s
            """,
            (table_name,),
        )
        return [(name, dtype) for name, dtype in cursor.fetchall()]
    finally:
        cursor.close()


def infer_users_name_expr(connection, users_table: str) -> str:
    override_col = os.getenv("SALESYNC_USERS_NAME_COLUMN")
    if override_col:
        return f"u.`{override_col}`"

    columns = dict(get_columns_for_table(connection, users_table))
    candidates = ["name", "full_name", "display_name", "username", "user_name"]
    for col in candidates:
        if col in columns:
            return f"u.`{col}`"

    if "first_name" in columns and "last_name" in columns:
        return "CONCAT_WS(' ', u.`first_name`, u.`last_name`)"

    for col, dtype in columns.items():
        if "name" in col.lower():
            return f"u.`{col}`"

    users_pk = get_primary_key_column(connection, users_table) or "id"
    return f"CAST(u.`{users_pk}` AS CHAR)"


def infer_checkins_columns(connection, checkins_table: str, users_table: str) -> tuple[str, str, str]:
    checkins_pk = os.getenv("SALESYNC_CHECKINS_PK") or get_primary_key_column(connection, checkins_table) or "id"

    user_fk_override = os.getenv("SALESYNC_CHECKINS_USER_ID_COLUMN")
    if user_fk_override:
        checkins_user_id_col = user_fk_override
    else:
        cursor = connection.cursor()
        try:
            cursor.execute(
                """
                SELECT COLUMN_NAME
                FROM information_schema.key_column_usage
                WHERE table_schema = DATABASE()
                  AND table_name = %s
                  AND REFERENCED_TABLE_NAME = %s
                ORDER BY ORDINAL_POSITION
                LIMIT 1
                """,
                (checkins_table, users_table),
            )
            row = cursor.fetchone()
            if row:
                checkins_user_id_col = row[0]
            else:
                cols = [c for c, _ in get_columns_for_table(connection, checkins_table)]
                checkins_user_id_col = next(
                    (c for c in cols if "user" in c.lower() or c.lower().endswith("_id")), cols[0]
                )
        finally:
            cursor.close()

    time_override = os.getenv("SALESYNC_CHECKINS_TIME_COLUMN")
    if time_override:
        checkins_time_col = time_override
    else:
        columns = get_columns_for_table(connection, checkins_table)
        datetime_like = [(name, dtype) for name, dtype in columns if dtype.lower() in ("datetime", "timestamp", "date")]
        name_priority = [
            "checkin_time",
            "check_in_time",
            "visited_at",
            "visit_time",
            "visit_date",
            "checkin_at",
            "created_at",
            "created_on",
            "time_in",
            "time",
            "date",
        ]
        checkins_time_col = next(
            (name for name, dtype in datetime_like if name.lower() in name_priority),
            datetime_like[0][0] if datetime_like else "checkin_time",
        )

    return checkins_pk, checkins_user_id_col, checkins_time_col


def get_earliest_visit_date(connection, checkins_table: str, checkins_time_col: str) -> Optional[date]:
    cursor = connection.cursor()
    try:
        query = f"SELECT DATE(MIN(`{checkins_time_col}`)) FROM `{checkins_table}`"
        cursor.execute(query)
        row = cursor.fetchone()
        if not row:
            return None
        value = row[0]
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if value is None:
            return None
        return date.fromisoformat(str(value))
    finally:
        cursor.close()


def fetch_daily_visits(start_date: Optional[date], end_date: Optional[date]):
    connection = mysql.connector.connect(**DATABASE_CONFIG)
    try:
        users_table = resolve_table_name(connection, [os.getenv("SALESYNC_TABLE_USERS"), "users", "user"]) or "users"
        checkins_table = resolve_table_name(connection, [os.getenv("SALESYNC_TABLE_CHECKINS"), "checkins", "visits"]) or "checkins"
        vr_table_resolved = resolve_table_name(
            connection, [os.getenv("SALESYNC_TABLE_VISIT_RESPONSE"), "visit_response", "visit_responses"]
        )
        include_vr = bool(vr_table_resolved)
        visit_response_table = vr_table_resolved or "visit_response"

        users_pk = os.getenv("SALESYNC_USERS_PK") or get_primary_key_column(connection, users_table) or "id"
        users_name_expr = infer_users_name_expr(connection, users_table)
        checkins_pk, checkins_user_id_col, checkins_time_col = infer_checkins_columns(connection, checkins_table, users_table)
        vr_checkin_id_col = os.getenv("SALESYNC_VISIT_RESPONSE_CHECKIN_ID_COLUMN", "checkin_id")

        if start_date is None:
            start_date = get_earliest_visit_date(connection, checkins_table, checkins_time_col)

        cursor = connection.cursor(dictionary=True)
        query, params = build_query(
            start_date,
            end_date,
            users_table=users_table,
            users_pk=users_pk,
            users_name_expr=users_name_expr,
            checkins_table=checkins_table,
            checkins_pk=checkins_pk,
            checkins_user_id_col=checkins_user_id_col,
            checkins_time_col=checkins_time_col,
            include_visit_response=include_vr,
            visit_response_table=visit_response_table,
            vr_checkin_id_col=vr_checkin_id_col,
        )
        cursor.execute(query, params)
        for row in cursor:
            yield {
                "user_name": row["user_name"],
                "date": row["visit_date"],
                "total_visits": int(row["total_visits"]) if row["total_visits"] is not None else 0,
            }
    finally:
        cursor.close()
        connection.close()


def write_pivot_csv(rows, output_path: str):
    user_names: set[str] = set()
    date_to_user_counts: dict[date, dict[str, int]] = {}

    for row in rows:
        user_name = str(row["user_name"]) if row["user_name"] is not None else ""
        raw_date = row["date"]
        if isinstance(raw_date, datetime):
            d = raw_date.date()
        elif isinstance(raw_date, date):
            d = raw_date
        else:
            try:
                d = date.fromisoformat(str(raw_date))
            except Exception:
                continue

        user_names.add(user_name)
        if d not in date_to_user_counts:
            date_to_user_counts[d] = {}
        date_to_user_counts[d][user_name] = int(row.get("total_visits", 0) or 0)

    sorted_users = sorted(user_names)
    header = ["Date", *sorted_users]
    ordered_dates = sorted(date_to_user_counts.keys())
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, mode="w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(header)
        for d in ordered_dates:
            display_date = d.strftime("%d-%b-%y")
            row_values = [display_date]
            counts_for_day = date_to_user_counts.get(d, {})
            for user in sorted_users:
                row_values.append(counts_for_day.get(user, 0))
            writer.writerow(row_values)


def write_pivot_xlsx(rows, output_path: str):
    user_names = sorted({r["user_name"] for r in rows})
    date_to_user_counts = {}
    for r in rows:
        d = r["date"]
        if isinstance(d, datetime):
            d = d.date()
        date_to_user_counts.setdefault(d, {u: 0 for u in user_names})
        date_to_user_counts[d][r["user_name"]] = r["total_visits"]

    ordered_dates = sorted(date_to_user_counts.keys())
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Visits"

    headers = ["Date"] + user_names
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for d in ordered_dates:
        row = [d.strftime("%d-%b-%y")] + [date_to_user_counts[d].get(u, 0) for u in user_names]
        ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    xlsx_path = output_path.replace(".csv", ".xlsx")
    wb.save(xlsx_path)
    print(f" Excel file saved: {xlsx_path}")


def generate_xlsx_bytes(start_date=None, end_date=None):
    """Generate the daily visits Excel file as bytes for API download."""
    # Parse dates
    if start_date:
        try:
            start_date_obj = date.fromisoformat(start_date)
        except Exception:
            start_date_obj = None
    else:
        start_date_obj = None
    if end_date:
        try:
            end_date_obj = date.fromisoformat(end_date)
        except Exception:
            end_date_obj = None
    else:
        end_date_obj = None
    # Fetch and pivot data
    rows = list(fetch_daily_visits(start_date_obj, end_date_obj))
    user_names = sorted({r["user_name"] for r in rows})
    date_to_user_counts = {}
    for r in rows:
        d = r["date"]
        if isinstance(d, datetime):
            d = d.date()
        date_to_user_counts.setdefault(d, {u: 0 for u in user_names})
        date_to_user_counts[d][r["user_name"]] = r["total_visits"]
    ordered_dates = sorted(date_to_user_counts.keys())
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Visits"
    headers = ["Date"] + user_names
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for d in ordered_dates:
        row = [d.strftime("%d-%b-%y")] + [date_to_user_counts[d].get(u, 0) for u in user_names]
        ws.append(row)
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def generate_csv_bytes(start_date=None, end_date=None):
    """Generate the daily visits CSV file as bytes for API download."""
    # Parse dates
    if start_date:
        try:
            start_date_obj = date.fromisoformat(start_date)
        except Exception:
            start_date_obj = None
    else:
        start_date_obj = None
    if end_date:
        try:
            end_date_obj = date.fromisoformat(end_date)
        except Exception:
            end_date_obj = None
    else:
        end_date_obj = None
    # Fetch and pivot data
    rows = list(fetch_daily_visits(start_date_obj, end_date_obj))
    user_names = sorted({r["user_name"] for r in rows})
    date_to_user_counts = {}
    for r in rows:
        d = r["date"]
        if isinstance(d, datetime):
            d = d.date()
        date_to_user_counts.setdefault(d, {u: 0 for u in user_names})
        date_to_user_counts[d][r["user_name"]] = r["total_visits"]
    ordered_dates = sorted(date_to_user_counts.keys())
    from io import StringIO
    output = StringIO()
    import csv
    header = ["Date"] + user_names
    writer = csv.writer(output)
    writer.writerow(header)
    for d in ordered_dates:
        display_date = d.strftime("%d-%b-%y")
        row_values = [display_date]
        counts_for_day = date_to_user_counts.get(d, {})
        for user in user_names:
            row_values.append(counts_for_day.get(user, 0))
        writer.writerow(row_values)
    return output.getvalue().encode("utf-8")


def main():
    load_dotenv()
    start_date = parse_date(sys.argv[1]) if len(sys.argv) > 1 and sys.argv[1] else None
    end_date = parse_date(sys.argv[2]) if len(sys.argv) > 2 and sys.argv[2] else None
    output_csv = sys.argv[3] if len(sys.argv) > 3 else os.path.join("reports", "daily_visits.csv")

    rows = list(fetch_daily_visits(start_date, end_date))
    write_pivot_csv(rows, output_csv)
    write_pivot_xlsx(rows, output_csv)
    print(f"\n Wrote {len(rows)} rows to {output_csv}\n")


if __name__ == "__main__":
    main()